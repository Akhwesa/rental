from django.shortcuts import render, redirect
from django.contrib.auth.models import Group
from django.utils import timezone
from datetime import datetime
from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Payment, Unit,Feed, Apartment, status
from django.db.models import Count
from user.models import Tenant, Landlord, User, Agent


# Create your report views here.
def receipt_excel(request):

    agent = Agent.objects.get(name = request.user)
    apartments = Apartment.objects.filter(agent=agent)
    unit = Unit.objects.filter(apartment__in=apartments)
    items = Payment.objects.filter(accountReference__in=unit.values('door_no'))

    # Create a new workbook
    workbook = Workbook()

    # Get the active worksheet
    worksheet = workbook.active

    # Define the column headers
    worksheet['A1'] = 'accountReference'
    worksheet['B1'] = 'paidAmount'
    worksheet['C1'] = 'paymentDate'
    worksheet['D1'] = 'fullName'
    worksheet['E1'] = 'transactionId'
    worksheet['F1'] = 'phoneNumber'
    worksheet['G1'] = 'invoiceName'
    worksheet['H1'] = 'externalReference'

    # Fill in the rows with data
    for i, item in enumerate(items):
        worksheet.cell(row=i+2, column=1, value=item.accountReference)
        worksheet.cell(row=i+2, column=2, value=item.paidAmount)
        worksheet.cell(row=i+2, column=3, value=item.paymentDate)
        worksheet.cell(row=i+2, column=5, value=item.transactionId)
        worksheet.cell(row=i+2, column=6, value=item.phoneNumber)
        worksheet.cell(row=i+2, column=4, value=item.fullName)
        worksheet.cell(row=i+2, column=7, value=item.invoiceName)
        worksheet.cell(row=i+2, column=8, value=item.externalReference)

    # Create a response object with the workbook as the content
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="agent_receipt.xlsx"'
    workbook.save(response)

    return response
def landlord_receipt(request):

    landlord = Landlord.objects.get(name = request.user)
    apartments = Apartment.objects.filter(landlord=landlord)
    unit = Unit.objects.filter(apartment__in=apartments)
    items = Payment.objects.filter(accountReference__in=unit.values('door_no'))

    # Create a new workbook
    workbook = Workbook()

    # Get the active worksheet
    worksheet = workbook.active

    # Define the column headers
    worksheet['A1'] = 'accountReference'
    worksheet['B1'] = 'paidAmount'
    worksheet['C1'] = 'paymentDate'
    worksheet['D1'] = 'fullName'
    worksheet['E1'] = 'transactionId'
    worksheet['F1'] = 'phoneNumber'
    worksheet['G1'] = 'invoiceName'
    worksheet['H1'] = 'externalReference'

    # Fill in the rows with data
    for i, item in enumerate(items):
        worksheet.cell(row=i+2, column=1, value=item.accountReference)
        worksheet.cell(row=i+2, column=2, value=item.paidAmount)
        worksheet.cell(row=i+2, column=3, value=item.paymentDate)
        worksheet.cell(row=i+2, column=5, value=item.transactionId)
        worksheet.cell(row=i+2, column=6, value=item.phoneNumber)
        worksheet.cell(row=i+2, column=4, value=item.fullName)
        worksheet.cell(row=i+2, column=7, value=item.invoiceName)
        worksheet.cell(row=i+2, column=8, value=item.externalReference)

    # Create a response object with the workbook as the content
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="landlord_receipt.xlsx"'
    workbook.save(response)

    return response

def monthly_receipt(request):
    if request.user.groups.filter(name='Landlord').exists(): 
        current_month = timezone.now().month

        landlord = Landlord.objects.get(name = request.user)
        apartments = Apartment.objects.filter(landlord=landlord)
        unit = Unit.objects.filter(apartment__in=apartments)
        items = Payment.objects.filter(accountReference__in=unit.values('door_no')).filter(paymentDate__month=current_month)

        # Create a new workbook
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Define the column headers
        worksheet['A1'] = 'accountReference'
        worksheet['B1'] = 'paidAmount'
        worksheet['C1'] = 'paymentDate'
        worksheet['D1'] = 'transactionId'
        worksheet['E1'] = 'Phone Number'
        worksheet['F1'] = 'Name'
        worksheet['G1'] = 'invoiceName'
        worksheet['H1'] = 'externalReference'

        # Fill in the rows with data
        for i, item in enumerate(items):
            worksheet.cell(row=i+2, column=1, value=item.accountReference)
            worksheet.cell(row=i+2, column=2, value=item.paidAmount)
            worksheet.cell(row=i+2, column=3, value=item.paymentDate)
            worksheet.cell(row=i+2, column=4, value=item.transactionId)
            worksheet.cell(row=i+2, column=5, value=item.phoneNumber)
            worksheet.cell(row=i+2, column=6, value=item.fullName)
            worksheet.cell(row=i+2, column=7, value=item.invoiceName)
            worksheet.cell(row=i+2, column=8, value=item.externalReference)

        # Create a response object with the workbook as the content
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="landlord_receipt.xlsx"'
        workbook.save(response)

        return response
    elif request.user.groups.filter(name='Agent').exists(): 
        current_month = timezone.now().month

        agent = Agent.objects.get(name = request.user)
        apartments = Apartment.objects.filter(agent=agent)
        unit = Unit.objects.filter(apartment__in=apartments)
        items = Payment.objects.filter(accountReference__in=unit.values('door_no')).filter(paymentDate__month=current_month)

        # Create a new workbook
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Define the column headers
        worksheet['A1'] = 'accountReference'
        worksheet['B1'] = 'paidAmount'
        worksheet['C1'] = 'paymentDate'
        worksheet['D1'] = 'transactionId'
        worksheet['E1'] = 'Phone Number'
        worksheet['F1'] = 'Name'
        worksheet['G1'] = 'invoiceName'
        worksheet['H1'] = 'externalReference'

        # Fill in the rows with data
        for i, item in enumerate(items):
            worksheet.cell(row=i+2, column=1, value=item.accountReference)
            worksheet.cell(row=i+2, column=2, value=item.paidAmount)
            worksheet.cell(row=i+2, column=3, value=item.paymentDate)
            worksheet.cell(row=i+2, column=4, value=item.transactionId)
            worksheet.cell(row=i+2, column=5, value=item.phoneNumber)
            worksheet.cell(row=i+2, column=6, value=item.fullName)
            worksheet.cell(row=i+2, column=7, value=item.invoiceName)
            worksheet.cell(row=i+2, column=8, value=item.externalReference)

        # Create a response object with the workbook as the content
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="landlord_receipt.xlsx"'
        workbook.save(response)

        return response

def tenant_arrear(request):
    if request.user.groups.filter(name='Landlord').exists(): 
        landlord = Landlord.objects.get(name = request.user)
        tenants = Tenant.objects.none()
        apartments = Apartment.objects.filter(landlord=landlord)
        unit = Unit.objects.filter(apartment__in=apartments)
        tenant_units = unit.filter(tenant__isnull=False)
        arr_units = tenant_units.exclude(Rent_status = 'Paid')
        items = arr_units.select_related('tenant')


        for item in items:
            tenants |= Tenant.objects.filter(pk = item.tenant.pk)
        

        # Create a new workbook
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Define the column headers
        worksheet['A1'] = 'name'
        worksheet['B1'] = 'tenant id'
        worksheet['C1'] = 'phone'
        worksheet['D1'] = 'mail'

        # Fill in the rows with data
        for i, tnt in enumerate(tenants):
            worksheet.cell(row=i+2, column=1, value=tnt.name)
            worksheet.cell(row=i+2, column=2, value=tnt.te_id)
            worksheet.cell(row=i+2, column=3, value=tnt.phone)
            worksheet.cell(row=i+2, column=4, value=tnt.mail)

        # Create a response object with the workbook as the content
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="landlord_receipt.xlsx"'
        workbook.save(response)

        return response
    elif request.user.groups.filter(name='Agent').exists(): 
        agent = Agent.objects.get(name = request.user)
        tenants = Tenant.objects.none()
        apartments = Apartment.objects.filter(agent=agent)
        unit = Unit.objects.filter(apartment__in=apartments)
        tenant_units = unit.filter(tenant__isnull=False)
        arr_units = tenant_units.exclude(Rent_status = 'Paid')
        items = arr_units.select_related('tenant')


        for item in items:
            tenants |= Tenant.objects.filter(pk = item.tenant.pk)
        

        # Create a new workbook
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Define the column headers
        worksheet['A1'] = 'name'
        worksheet['B1'] = 'tenant id'
        worksheet['C1'] = 'phone'
        worksheet['D1'] = 'mail'

        # Fill in the rows with data
        for i, tnt in enumerate(tenants):
            worksheet.cell(row=i+2, column=1, value=tnt.name)
            worksheet.cell(row=i+2, column=2, value=tnt.te_id)
            worksheet.cell(row=i+2, column=3, value=tnt.phone)
            worksheet.cell(row=i+2, column=4, value=tnt.mail)

        # Create a response object with the workbook as the content
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="landlord_receipt.xlsx"'
        workbook.save(response)

        return response

def landlord_units(request):
    if request.user.groups.filter(name='Landlord').exists():

        landlord = Landlord.objects.get(name = request.user)
        apartments = Apartment.objects.filter(landlord=landlord)
        items = Unit.objects.filter(apartment__in=apartments)

        # Create a new workbook
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Define the column headers
        worksheet['A1'] = 'apartment'
        worksheet['B1'] = 'door no'
        worksheet['C1'] = 'Type'
        worksheet['D1'] = 'rent'
        worksheet['E1'] = 'rent Status'
        worksheet['F1'] = 'House Status'
        worksheet['G1'] = 'Tenant'

        # Fill in the rows with data
        for i, item in enumerate(items):
            worksheet.cell(row=i+2, column=1, value=item.apartment.name)
            worksheet.cell(row=i+2, column=2, value=item.door_no)
            worksheet.cell(row=i+2, column=3, value=item.type)
            worksheet.cell(row=i+2, column=4, value=item.rent)
            worksheet.cell(row=i+2, column=5, value=item.Rent_status)
            worksheet.cell(row=i+2, column=6, value=item.hse_status)
            worksheet.cell(row=i+2, column=7, value=item.tenant.name if item.tenant else 'null')

        # Create a response object with the workbook as the content
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="House Units.xlsx"'
        workbook.save(response)

        return response
    elif request.user.groups.filter(name='Agent').exists():
        agent = Agent.objects.get(name = request.user)
        apartments = Apartment.objects.filter(agent=agent)
        items = Unit.objects.filter(apartment__in=apartments)

        # Create a new workbook
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Define the column headers
        worksheet['A1'] = 'apartment'
        worksheet['B1'] = 'door no'
        worksheet['C1'] = 'Type'
        worksheet['D1'] = 'rent'
        worksheet['E1'] = 'rent Status'
        worksheet['F1'] = 'House Status'
        worksheet['G1'] = 'Tenant'

        # Fill in the rows with data
        for i, item in enumerate(items):
            worksheet.cell(row=i+2, column=1, value=item.apartment.name)
            worksheet.cell(row=i+2, column=2, value=item.door_no)
            worksheet.cell(row=i+2, column=3, value=item.type)
            worksheet.cell(row=i+2, column=4, value=item.rent)
            worksheet.cell(row=i+2, column=5, value=item.Rent_status)
            worksheet.cell(row=i+2, column=6, value=item.hse_status)
            worksheet.cell(row=i+2, column=7, value=item.tenant.name if item.tenant else 'null')

        # Create a response object with the workbook as the content
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="House Units.xlsx"'
        workbook.save(response)

        return response
